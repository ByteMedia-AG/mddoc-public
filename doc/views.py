import csv
import datetime
import difflib
import io
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from decimal import Decimal
from io import BytesIO
from itertools import groupby

import openpyxl
from bs4 import BeautifulSoup
from django.conf import settings
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.contenttypes.models import ContentType
from django.db import connection
from django.db.models import Min, Max
from django.db.models import Subquery
from django.db.models.functions import Lower
from django.db.utils import OperationalError
from django.http import HttpResponseRedirect, HttpResponse
from django.template.response import TemplateResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.dateparse import parse_datetime
from django.utils.timezone import localtime
from openpyxl.styles import Alignment, Border, Side
from taggit.models import Tag, TaggedItem

from .forms import DocForm, TimeRecordForm, CleanupForm
from .markdown import md2html
from .models import Doc, TimeRecord
from .models import File
from .utils import extract_selected_info_from_url
from django.core.files import File as DjangoFile
import traceback
from pathlib import Path
from django.db.models import Count


def get_tags_grouped(tags):
    """
    Von jedem Anfangsbuchstaben werde jeweils nur die 5 häufigsten zur Anzeige gebracht.
    """
    try:
        grouped = []
        tags_sorted = sorted(tags, key=lambda t: (t.slug[0].upper(), -getattr(t, 'num_times', 0)))
        for first_letter, group in groupby(tags_sorted, key=lambda t: t.slug[0].upper()):
            group_list = list(group)
            top_tags = sorted(group_list[:5], key=lambda t: t.slug.lower())  # alphabetisch sortieren nach Begrenzung
            for i, tag in enumerate(top_tags):
                tag.is_last_in_group = (i == len(top_tags) - 1)
            grouped.append((first_letter, top_tags))
        return grouped
    except Exception as e:
        return []


def parse_valid_date(value, fallback):
    try:
        parsed = parse_date(value)
        return parsed or fallback
    except Exception:
        return fallback


def flatten_settings(data, prefix=''):
    flat = {}
    if isinstance(data, dict):
        for key, value in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            flat.update(flatten_settings(value, full_key))
    else:
        flat[prefix] = data
    return flat


@login_required
@permission_required("doc.search")
def docs(request):
    """List of documents."""

    tags = None
    error = False
    num_matches = 0

    if request.GET:
        params = request.GET.dict()
        request.session['docs_params'] = params
    else:
        params = request.session.get('docs_params', {})

    asc = bool(int(params.get('asc', 0)))
    oby = params.get('orderby', 'created')
    oby2 = oby
    if oby == 'updated':
        oby = 'updated_at'
    elif oby == 'title':
        oby = 'title'
    elif oby == 'time':
        oby = 'time'
    else:
        oby = 'created_at'
    if not asc:
        oby = f'-{oby}'

    search_string = params.get('search', False)
    try:
        if search_string:
            entities = Doc.objects.match(search_string)
        else:
            entities = Doc.objects.all()
        if not params.get('ipr', False):
            # Do not include previous revisions.
            entities = entities.exclude(successor__isnull=False)
        if not params.get('idr', False):
            # Do not include deleted resources.
            entities = entities.exclude(successor__isnull=True, deleted_at__isnull=False)
        if not params.get('iar', False):
            # Do not include archived resources.
            entities = entities.exclude(is_archived=True)
        if params.get('flg', False):
            entities = entities.filter(is_flagged=True)
        if params.get('ustr', False):
            entities = entities.filter(has_unsettled_tr=True)
        if params.get('cgte', False):
            entities = entities.filter(created_at__gte=timezone.make_aware(parse_datetime(params.get('cgte'))))
        if params.get('clte', False):
            entities = entities.filter(created_at__lte=timezone.make_aware(parse_datetime(params.get('clte'))))
        if params.get('ugte', False):
            entities = entities.filter(updated_at__gte=timezone.make_aware(parse_datetime(params.get('ugte'))))
        if params.get('ulte', False):
            entities = entities.filter(updated_at__lte=timezone.make_aware(parse_datetime(params.get('ulte'))))
        if params.get('tgte', False):
            entities = entities.filter(time__gte=timezone.make_aware(parse_datetime(params.get('tgte'))))
        if params.get('tlte', False):
            entities = entities.filter(time__lte=timezone.make_aware(parse_datetime(params.get('tlte'))))
        num_matches = entities.count()
        entities = entities.order_by(oby)[:100]
        entity_ids = [e.id for e in entities]
        available_tags = Doc.tags \
                             .most_common(extra_filters={'doc__in': entity_ids}) \
                             .filter(num_times__lt=len(entity_ids), num_times__gt=0) \
                             .order_by('-num_times', 'name')[:1000]
        if search_string:
            tags = [tag for tag in available_tags if tag.name.lower() not in search_string.lower()]
        else:
            tags = [tag for tag in available_tags]
    except OperationalError as oe:
        entities = Doc.objects.filter(deleted_at__isnull=True, is_archived=False).order_by('-id')[:10]
        error = oe
        print(oe)
    except Exception as e:
        print(e.__traceback__)

    if False:
        print("===================================")
        for query in connection.queries:
            time = float(query['time'])
            print(f"{time}  s: {query['sql']}")
            print("--------------------------------")

    return TemplateResponse(request, "docs.html", {
        'page_title': 'Search Resources',
        'entities': entities,
        'tags': get_tags_grouped(tags),
        'error': error,
        'asc': asc,
        'oby': oby2,
        'params': params,
        'num_matches': num_matches,
    })


@login_required
@permission_required("doc.add_time_record")
def time_add(request, **kwargs):
    """Add a new time record"""

    doc_id = kwargs.get('doc_id')

    if request.method != 'POST':
        return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))

    try:
        doc = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    trForm = TimeRecordForm(request.POST, user=request.user)
    if trForm.is_valid():
        time_record = trForm.save(commit=False)
        time_record.doc = doc
        time_record.save()
        doc.save()

    return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))


@login_required
@permission_required("doc.time_records_set_settled")
def time_records_set_settled(request, **kwargs):
    """
    Sets the selected time records to settled.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("docs"))

    success = False
    error = False

    try:
        trs = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user)
        for tr in trs:
            tr.settled_at = timezone.now()
            tr.save()
            tr.doc.save()
            success = [
                'The selected entries were set to settled.',
                'Please use the back button in the toolbar to return to the start page.',
            ]
    except Exception as e:
        error = [e]

    return TemplateResponse(request, "message.html", {
        'page_title': f"Set Time Records To Settled",
        'success': success,
        'error': error,
    })


@login_required
@permission_required("doc.time_records_set_deleted")
def time_records_set_deleted(request, **kwargs):
    """
    Sets the selected time records to deleted.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("docs"))

    success = False
    error = False

    try:
        trs = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user)
        for tr in trs:
            tr.deleted_at = timezone.now()
            tr.save()
            tr.doc.save()
            success = [
                'The selected entries have been deleted.',
                'Please use the back button in the toolbar to return to the start page.',
            ]
    except Exception as e:
        error = [e]

    return TemplateResponse(request, "message.html", {
        'page_title': f"Delete Time Records",
        'success': success,
        'error': error,
    })


@login_required
@permission_required("doc.time_analyse")
def time_analyse_download(request, **kwargs):
    """
    Returns a download in the specified format.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("docs"))

    rows = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user, ) \
        .order_by('date', 'doc__title', 'description') \
        .values_list('date', 'doc__title', 'description', 'time', named=True)

    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"timereport_{timestamp}"

    if request.POST.get('dcsv', default=False):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Item", "Task", "Time [h]"])
        for row in rows:
            writer.writerow([row.date, row.doc__title, row.description, row.time])
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response
    elif request.POST.get('dxlsx', default=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Time Report"
        headers = ["Date", "Item", "Task", "Time [h]"]
        ws.append(headers)

        # Freeze the first row and set it as title row
        ws.freeze_panes = "A2"
        ws.print_title_rows = '1:1'

        # Header in bold letters
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Add rows
        for row in rows:
            ws.append([row.date, row.doc__title, row.description, row.time])

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 11

        # Row height
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 31

        # Alignment of the columns A, B and C
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Allow text wrap in the columns B and C
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        # Alignment of the column D
        ws["D1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Number format and alignment of the column D
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        thin_border = Border(bottom=Side(style="thin", color="000000"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

        # Page properties
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.59
        ws.page_margins.right = 0.59
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response
    elif request.POST.get('dpdf', default=False):
        """The PDF format will not be implemented in the near future."""
        total = sum((Decimal(str(row.time)) for row in rows), Decimal('0.0'))

    response = HttpResponse("No format specified", content_type="text/plain")
    response["Content-Disposition"] = f'attachment; filename="{filename}.txt"'
    return response


@login_required
@permission_required("doc.time_analyse")
def time_analyse_preview(request, **kwargs):
    """
    Lists the selected time records.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("docs"))

    rows = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user, ) \
        .order_by('date', 'doc__title', 'description') \
        .values_list('date', 'doc__title', 'description', 'time', named=True)

    total = sum((Decimal(str(row.time)) for row in rows), Decimal('0.0'))

    return TemplateResponse(request, "tr_analyse_preview.html", {
        'page_title': f"Time Records Selection",
        'rows': rows,
        'total': total,
    })


@login_required
@permission_required("doc.bulk_update")
def bulk_update(request, **kwargs):
    """
    Allows to add/remove is_flagged, is_archived and tags.
    """

    success = []
    error = []

    show_exclusion_deleted_info = False
    show_initial_warning = False
    no_docs_left = False
    show_delete_warning = False

    if request.method == 'GET':
        doc_list = request.GET.getlist('doc') or []
        if len(doc_list) < 1:
            return HttpResponseRedirect(reverse("docs"))
        docs = Doc.objects.filter(
            successor__isnull=True,
            id__in=doc_list,
        ).order_by(Lower('title'))
        show_initial_warning = True
        if len(request.GET.getlist('doc') or []) > docs.count():
            show_exclusion_deleted_info = True

    if request.method == 'POST':
        docs = Doc.objects.filter(
            successor__isnull=True,
            id__in=request.POST.getlist('doc') or [],
        ).order_by(Lower('title'))

    if docs.count() < 1:
        no_docs_left = True

    if request.method == 'POST':
        try:
            for doc in docs:
                if request.POST.get('set-flag'):
                    doc.is_flagged = True
                    success.append(f'Doc {doc.id} is flagged.')
                if request.POST.get('unset-flag'):
                    doc.is_flagged = False
                    success.append(f'Doc {doc.id} is not flagged.')
                if request.POST.get('set-archived'):
                    doc.is_archived = True
                    doc.deleted_at = None
                    show_delete_warning = True
                    success.append(f'Doc {doc.id} is archived.')
                if request.POST.get('unset-archived'):
                    doc.is_archived = False
                    success.append(f'Doc {doc.id} is not archived.')
                if request.POST.get('set-deleted'):
                    doc.deleted_at = timezone.now()
                    doc.is_archived = False
                    show_delete_warning = True
                    success.append(f'Doc {doc.id} is marked as deleted.')
                if request.POST.get('unset-deleted'):
                    doc.deleted_at = None
                    success.append(f'Doc {doc.id} is not marked as deleted.')
                if request.POST.get('rm-tag'):
                    try:
                        tag = Tag.objects.get(id=request.POST.get('rm-tag'))
                        doc.tags.remove(tag)
                        success.append(f'Doc {doc.id} no longer has the tag "{tag.name}".')
                    except Tag.DoesNotExist:
                        error.append(f"Tag with ID {request.POST.get('rm-tag')} does not exist.")
                if request.POST.get('add-tag'):
                    tag_string = request.POST.get('add-tag')
                    tag_list = [tag.strip() for tag in tag_string.split(',') if tag.strip()]
                    for tag in tag_list:
                        doc.tags.add(tag)
                    success.append(f'Doc {doc.id} now has the tags "{tag_list}".')
                doc.tag = " ".join(sorted(doc.tags.slugs()))
                doc.save()
        except Exception as e:
            success = []
            error.append(e)

    tags = Doc.tags.filter(doc__in=docs).distinct().order_by('name')

    return TemplateResponse(request, "bulk_update.html", {
        'page_title': f"Bulk update",
        'docs': docs,
        'tags': tags,
        'show_exclusion_deleted_info': show_exclusion_deleted_info,
        'show_initial_warning': show_initial_warning,
        'no_docs_left': no_docs_left,
        'success': success,
        'error': error,
        'show_delete_warning': show_delete_warning,
    })


@login_required
@permission_required("doc.time_analyse")
def time_analyse_selection(request, **kwargs):
    """
    Lists all unsettled time record entries of the given docs and allows to discriminate
    to a certain period of time and to select a set of entries to be previewed, downloaded
    and set to billed.
    """

    entities = TimeRecord.objects.filter(
        deleted_at__isnull=True,
        settled_at__isnull=True,
        doc_id__in=request.GET.getlist('doc') or [],
        user=request.user,
    )

    date_range = entities.aggregate(
        min_date=Min('date'),
        max_date=Max('date')
    )

    dgte = parse_valid_date(request.GET.get('dgte'), date_range['min_date'])
    dlte = parse_valid_date(request.GET.get('dlte'), date_range['max_date'])
    if dgte and dlte and dgte > dlte:
        dgte, dlte = dlte, dgte

    entities = entities.filter(date__gte=dgte, date__lte=dlte).order_by('date', 'doc__title', 'description')

    return TemplateResponse(request, "tr_analyse_selection.html", {
        'page_title': f"Time Records Selection",
        'min_date': date_range['min_date'],
        'max_date': date_range['max_date'],
        'dgte': dgte,
        'dlte': dlte,
        'docs': request.GET.getlist('doc') or [],
        'entities': entities,
    })


@login_required
@permission_required("doc.add")
def doc_add_short(request, **kwargs):
    """
    Shorter path to add a document. Only the description and optionally one
    or more tags are presented in the modal form. Other necessary information
    is added automatically to the entry.
    """

    if request.method != 'POST':
        return HttpResponseRedirect(reverse("docs"))

    description = request.POST.get('description', '').strip()
    tags = request.POST.get('tags', '').strip()

    if not description:
        return HttpResponseRedirect(reverse("docs"))

    now = timezone.now()
    local_now = localtime(now)
    datetime_string = local_now.strftime("%Y-%m-%d %H:%M %Z")

    d: Doc = Doc()
    d.title = f'Log - {datetime_string}'
    d.description = description
    d.time = now
    d.is_markdown = False
    d.is_archived = True
    d.is_flagged = False
    d.save()

    tag_list = [tag.strip() for tag in tags.split(',') if tag.strip()]
    for tag in tag_list:
        d.tags.add(tag)
    d.tags.add('log-entry')
    d.tag = " ".join(list(d.tags.slugs()))
    d.save()

    return HttpResponseRedirect(reverse("docs"))


@login_required
@permission_required("doc.add")
def doc_add(request, **kwargs):
    """Add a new document"""

    if request.method == 'POST':
        form = DocForm(request.POST, request.FILES)
        if form.is_valid():
            entity = form.save()
            for uploaded_file in request.FILES.getlist("upload"):
                file_obj = File.objects.create(file=uploaded_file, name=uploaded_file.name)
                entity.files.add(file_obj)
            entity.tag = " ".join(list(entity.tags.slugs()))
            entity.save()
            return HttpResponseRedirect(reverse("doc__detail", args=(entity.id,)))
    else:
        form = DocForm()

    return TemplateResponse(request, "doc-edit.html", {
        'page_title': "Add new document",
        'form': form,
    })


@login_required
@permission_required("doc.extract")
def doc_extract(request, **kwargs):
    """Extracts the text of the given URI."""

    doc_id = kwargs.get('id')
    error = None

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    if not entity.uri:
        return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))

    extracted_data = extract_selected_info_from_url(entity.uri)

    title_url = (entity.uri[:50] + '...') if len(entity.uri) > 50 else entity.uri

    return TemplateResponse(request, "doc_extract_text.html", {
        'page_title': f"Extract {title_url}",
        'error': error,
        'data': extracted_data,
    })


@login_required
@permission_required("doc.delete")
def doc_delete(request, **kwargs):
    """Sets the attribute doc.deleted_at to now()."""

    doc_id = kwargs.get('id')

    if request.method != 'POST':
        # In case of a GET request, a redirect to the corresponding details view is done.
        return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    if request.method == 'POST':
        entity.remove()
        return HttpResponseRedirect(reverse("docs"))


@login_required
@permission_required("doc.restore")
def doc_restore(request, **kwargs):
    """Either undeleted the document or copies the values of the select entry to the current representation."""

    doc_id = kwargs.get('id')

    if request.method != 'POST':
        # In case of a GET request, a redirect to the corresponding details view is done.
        return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    if entity.successor:
        entity.restore_revision()
        return HttpResponseRedirect(reverse("doc__detail", args=(entity.successor.id,)))
    else:
        # The deleted attribute has to be set to None
        entity.deleted_at = None
        entity.save()
        return HttpResponseRedirect(reverse("doc__detail", args=(entity.id,)))


@login_required
@permission_required("doc.edit")
def doc_edit(request, **kwargs):
    """Edit a document"""

    doc_id = kwargs.get('id')

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    show_revision_warning = False
    if request.method == 'GET' and not entity.is_archived:
        show_revision_warning = True

    if request.method == 'POST':
        form = DocForm(request.POST, request.FILES, instance=entity)
        print(request.FILES)
        print(request.FILES.getlist("upload"))
        print(form.errors)
        if form.is_valid():
            entity.make_revision()
            entity = form.save()
            for uploaded_file in request.FILES.getlist("upload"):
                file_obj = File.objects.create(file=uploaded_file, name=uploaded_file.name)
                entity.files.add(file_obj)
            entity.is_archived = False
            entity.tag = " ".join(list(entity.tags.slugs()))
            delete_ids = form.cleaned_data.get("delete_files", [])
            if delete_ids:
                entity.files.remove(*File.objects.filter(id__in=delete_ids))
            entity.save()
            return HttpResponseRedirect(reverse("doc__detail", args=(entity.id,)))
    else:
        form = DocForm(instance=entity)
    return TemplateResponse(request, "doc-edit.html", {
        'page_title': f"{entity.title}",
        'form': form,
        'show_revision_warning': show_revision_warning,
        'doc_id': doc_id,
    })


@login_required
@permission_required("doc.add_log")
def doc_add_log(request, **kwargs):
    """Adds a log entry."""

    doc_id = kwargs.get('id')

    log_entry = request.POST.get('log_entry', default='').strip()
    if log_entry != '':
        log_entry = BeautifulSoup(log_entry, features="html.parser").get_text()

    if request.method != 'POST' or log_entry == '':
        return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    now = datetime.datetime.now().replace(microsecond=0).isoformat()
    if entity.log:
        log_entry = f"<!-- {now} -->\n{log_entry}\n{entity.log}"
    else:
        log_entry = f"<!-- {now} -->\n{log_entry}"
    entity.log = log_entry
    entity.save()

    return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))


@login_required
@permission_required("doc.archive")
def doc_toggle_archive(request, **kwargs):
    """Set is_archived to True if it was False and the other way around."""

    doc_id = kwargs.get('id')

    try:
        entity = Doc.objects.get(pk=doc_id)
        if entity.is_archived:
            entity.is_archived = False
        else:
            entity.is_archived = True
        entity.save()
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))


@login_required
@permission_required("doc.flag")
def doc_toggle_flag(request, **kwargs):
    """Set is_flagged to True if it was False and the other way around."""

    doc_id = kwargs.get('id')

    try:
        entity = Doc.objects.get(pk=doc_id)
        if entity.is_flagged:
            entity.is_flagged = False
        else:
            entity.is_flagged = True
        entity.save()
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    return HttpResponseRedirect(reverse("doc__detail", args=(doc_id,)))


@login_required
@permission_required("doc.view")
def doc(request, **kwargs):
    """Gets the details of the requested document."""

    doc_id = kwargs.get('id')

    try:
        entity = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    html = ""
    if entity.is_markdown:
        html = md2html(entity.text)

    diff = None
    if entity.successor:
        diff = difflib.unified_diff(str(entity.text).splitlines(), str(entity.successor.text).splitlines())
        diff = '\n'.join(diff)

    log = None
    if entity.log:
        pattern = re.compile(r'<!--\s*(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})\s*-->\s*')
        matches = list(pattern.finditer(entity.log))
        log = []
        for i, match in enumerate(matches):
            dt = datetime.datetime.fromisoformat(match.group(1))
            start = match.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(entity.log)
            log.append((dt, entity.log[start:end].strip()))

    return TemplateResponse(request, "doc.html", {
        'page_title': f"{entity.title}",
        'entity': entity,
        'doc_id': doc_id,
        'html': html,
        'tags': list(entity.tags.all().order_by('name')),
        'is_deleted': True if entity.deleted_at else False,
        'is_revised': True if entity.successor else False,
        'successor_id': entity.successor.id if entity.successor else None,
        'revisions': entity.predecessors.order_by('updated_at').reverse(),
        'diff': diff,
        'log': log,
        'files': entity.files.all().order_by('name'),
        'time_record_form': TimeRecordForm(),
        'time_records': list(entity.time_records.all().order_by('-date')) if entity.time_records.exists() else False,
    })


@login_required
@permission_required("doc.cleanup")
def cleanup_database(request, **kwargs):
    """Allows to clean up the database."""

    success = False
    error = False

    if request.method == 'GET':
        form = CleanupForm()

        return TemplateResponse(request, "cleanup.html", {
            'page_title': "Database cleanup",
            'form': form,
            'success': success,
            'error': error,
        })

    if not request.POST:
        return HttpResponseRedirect(reverse("docs"))

    form = CleanupForm(request.POST)
    if form.is_valid():
        do_commit = request.POST.get('do_commit', default=False)
        number_of_deleted_docs = 0
        number_of_deleted_revisions = 0
        number_of_deleted_archived_revisions = 0
        number_of_deleted_tr = 0
        number_of_settled_tr = 0
        number_of_undeleted_tr = 0
        number_of_unsettled_tr = 0
        number_of_removed_tags = 0
        if request.POST.get('remove_deleted_docs', default=False):
            deleted_docs = Doc.objects.filter(deleted_at__isnull=False, successor__isnull=True)
            if form.cleaned_data['from_date']:
                deleted_docs = deleted_docs.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                deleted_docs = deleted_docs.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_deleted_docs = deleted_docs.count()
            if do_commit:
                for d in deleted_docs:
                    d.delete()
        if request.POST.get('remove_previous_revisions', default=False):
            revisions = Doc.objects.filter(deleted_at__isnull=False, successor__isnull=False)
            if form.cleaned_data['from_date']:
                revisions = revisions.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                revisions = revisions.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_deleted_revisions = revisions.count()
            if do_commit:
                for d in revisions:
                    d.delete()
        if request.POST.get('remove_archived_revisions', default=False):
            archived_revisions = Doc.objects.filter(deleted_at__isnull=False, successor__is_archived=True)
            if form.cleaned_data['from_date']:
                archived_revisions = archived_revisions.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                archived_revisions = archived_revisions.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_deleted_archived_revisions = archived_revisions.count()
            if do_commit:
                for d in archived_revisions:
                    d.delete()
        if request.POST.get('remove_unused_tags', default=False):
            doc_ct = ContentType.objects.get_for_model(Doc)
            live_doc_ids = Subquery(Doc.objects.values('id'))
            orphaned_tagged_items = TaggedItem.objects.filter(
                content_type=doc_ct
            ).exclude(object_id__in=live_doc_ids)
            number_of_removed_tags = orphaned_tagged_items.count()
            if do_commit:
                orphaned_tagged_items.delete()
        if request.POST.get('remove_deleted_timerecords', default=False):
            deleted_tr = TimeRecord.objects.filter(deleted_at__isnull=False)
            if form.cleaned_data['from_date']:
                deleted_tr = deleted_tr.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                deleted_tr = deleted_tr.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_deleted_tr = deleted_tr.count()
            if do_commit:
                for d in deleted_tr:
                    d.delete()
        if request.POST.get('remove_settled_timerecords', default=False):
            settled_tr = TimeRecord.objects.filter(settled_at__isnull=False)
            if form.cleaned_data['from_date']:
                settled_tr = settled_tr.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                settled_tr = settled_tr.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_settled_tr = settled_tr.count()
            if do_commit:
                for d in settled_tr:
                    d.delete()
        if request.POST.get('undo_delete_timerecords', default=False):
            restore_tr = TimeRecord.objects.filter(deleted_at__isnull=False)
            if form.cleaned_data['from_date']:
                restore_tr = restore_tr.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                restore_tr = restore_tr.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_undeleted_tr = restore_tr.count()
            if do_commit:
                for tr in restore_tr:
                    tr.deleted_at = None
                    tr.save()
                    tr.doc.save()
        if request.POST.get('undo_settle_timerecords', default=False):
            unsettle_tr = TimeRecord.objects.filter(settled_at__isnull=False)
            if form.cleaned_data['from_date']:
                unsettle_tr = unsettle_tr.filter(created_at__gte=form.cleaned_data['from_date'])
            if form.cleaned_data['to_date']:
                unsettle_tr = unsettle_tr.filter(created_at__lte=form.cleaned_data['to_date'])
            number_of_unsettled_tr = unsettle_tr.count()
            if do_commit:
                for tr in unsettle_tr:
                    tr.settled_at = None
                    tr.save()
                    tr.doc.save()
        success = [
            f'{number_of_deleted_docs} documents have been removed.',
            f'{number_of_deleted_revisions} revisions have been removed.',
            f'{number_of_deleted_archived_revisions} archived revisions have been removed.',
            f'{number_of_removed_tags} tags have been removed.',
            f'{number_of_deleted_tr} time records marked as deleted have been removed.',
            f'{number_of_settled_tr} time records marked as settled have been removed.',
            f'{number_of_undeleted_tr} time records marked as deleted have been restored.',
            f'{number_of_unsettled_tr} time records marked as settled have been restored.',
        ]

        context = {
            'page_title': "Database cleanup",
            'success': success,
            'error': error,
            'faked': True if not do_commit else False,
        }
        if not do_commit:
            context['form'] = form
        return TemplateResponse(request, "cleanup.html", context)
    else:
        return TemplateResponse(request, "cleanup.html", {
            'page_title': "Database cleanup",
            'form': form,
            'success': success,
            'error': error,
        })


@login_required
@permission_required("doc.dump_database")
def dump_database(request, **kwargs):
    """Does a dump of the database."""

    success = []
    error = []

    if request.method == 'POST':
        save_as_doc = request.POST.get('target_rmdoc', default=False)
        save_as_file = request.POST.get('target_filesystem', default=False)
        if not save_as_doc and not save_as_file:
            error.append('No storage location has been selected.')
        else:
            try:
                timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"sqlite_backup_{timestamp}.sql"
                db_path = settings.DATABASES['default']['NAME']

                if save_as_file and not save_as_doc:
                    backup_dir = os.path.join(settings.BASE_DIR, 'db_dumps')
                    os.makedirs(backup_dir, exist_ok=True)
                    backup_path = os.path.join(backup_dir, backup_filename)
                    with open(backup_path, 'w') as f:
                        subprocess.run(['sqlite3', db_path, '.dump'], check=True, stdout=f)
                    success.append(f"The database dump has been saved to: {backup_path}")
                else:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".sql") as tmp:
                        subprocess.run(['sqlite3', db_path, '.dump'], check=True, stdout=tmp)
                        tmp_path = tmp.name
                    # Create a Doc entry
                    with open(tmp_path, 'rb') as f:
                        now = timezone.now()
                        datetime_string = now.strftime("%Y-%m-%d %H:%M")
                        date_string = now.strftime("%Y-%m-%d")
                        doc = Doc.objects.create(
                            title=f"Database Backup {date_string}",
                            is_markdown=False,
                            is_archived=True,
                            time=now,
                            tag='db-dump',
                            description=f'This is an automatically created entry from a database backup. The dump was created {datetime_string}.',
                        )
                        doc.tags.add("db-dump")
                        django_file = DjangoFile(f)
                        django_file.name = Path(backup_filename).name
                        file_obj = File.objects.create(
                            file=django_file,
                            name=Path(backup_filename).name
                        )
                        doc.files.add(file_obj)
                        doc.save()
                        success.append(f"The database dump has been saved as document: {doc.title}")
                    if save_as_file:
                        # Move the file to the dir db_dumps
                        backup_dir = os.path.join(settings.BASE_DIR, 'db_dumps')
                        os.makedirs(backup_dir, exist_ok=True)
                        backup_path = os.path.join(backup_dir, backup_filename)
                        shutil.copy(tmp_path, backup_path)
                        success.append(f"The database dump has been saved to: {backup_path}")
                    os.remove(tmp_path)
            except Exception as e:
                traceback.print_exc()
                error.append(f"An exception occurred: {e}")

    return TemplateResponse(request, "database_dump.html", {
        'page_title': "Database dump",
        'success': success,
        'error': error,
    })


@login_required
@permission_required("doc.show_status")
def status(request, **kwargs):
    """Shows information about the data persisted."""

    docs = {}
    doc_time = {}
    trs = {}
    trs_time = {}
    file = {}

    docs['All'] = Doc.objects.all().count()
    docs['Current'] = Doc.objects.filter(
        deleted_at__isnull=True, is_archived=False).count()
    docs['Current, with unsettled time records'] = TimeRecord.objects.filter(
        settled_at__isnull=True, deleted_at__isnull=True,
        doc__deleted_at__isnull=True, doc__is_archived=False).values_list('doc_id', flat=True).distinct().count()
    docs['Current, with file attachment'] = Doc.objects.annotate(num_files=Count('files')).filter(
        deleted_at__isnull=True, is_archived=False, num_files__gt=0).count()
    docs['Archived'] = Doc.objects.filter(
        deleted_at__isnull=True, is_archived=True).count()
    docs['Versions'] = Doc.objects.filter(
        deleted_at__isnull=True, successor__isnull=False).count()
    docs['Revisions'] = Doc.objects.filter(
        deleted_at__isnull=False, successor__isnull=False).count()
    docs['Deleted'] = Doc.objects.filter(
        deleted_at__isnull=False, successor__isnull=True).count()

    doc_time['Youngest'] = Doc.objects.aggregate(Max('created_at'))['created_at__max']
    doc_time['Oldest'] = Doc.objects.aggregate(Min('created_at'))['created_at__min']

    trs['All'] = TimeRecord.objects.all().count()
    trs['Unsettled'] = TimeRecord.objects.filter(
        settled_at__isnull=True, deleted_at__isnull=True, doc__deleted_at__isnull=True, doc__is_archived=False).count()
    trs['Settled'] = TimeRecord.objects.filter(
        settled_at__isnull=False
    ).count()
    trs['Deleted'] = TimeRecord.objects.filter(
        deleted_at__isnull=False
    ).count()

    file['All'] = File.objects.all().count()
    file['Orphaned'] = File.objects.filter(docs__isnull=True).count()

    trs_time['Youngest'] = TimeRecord.objects.aggregate(Max('created_at'))['created_at__max']
    trs_time['Oldest'] = TimeRecord.objects.aggregate(Min('created_at'))['created_at__min']

    settings_dict = {
        k: getattr(settings, k)
        for k in dir(settings)
        if k.isupper()
    }
    flat_settings = flatten_settings(settings_dict)

    return TemplateResponse(request, "status.html", {
        'page_title': "Status",
        'docs': docs,
        'doc_time': doc_time,
        'trs': trs,
        'trs_time': trs_time,
        'file': file,
        'settings': flat_settings,
    })
