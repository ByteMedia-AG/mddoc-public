import csv
import io
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Min, Max
from django.http import HttpResponseRedirect, HttpResponse
from django.template.response import TemplateResponse
from django.urls import reverse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Side

from doc.forms import TimeRecordForm
from doc.models import Doc, TimeRecord
from doc.utils import parse_valid_date


@login_required
@permission_required("doc.add_time_record")
def time_add(request, **kwargs):
    """Add a new time record"""

    doc_id = kwargs.get('doc_id')

    if request.method != 'POST':
        return HttpResponseRedirect(reverse("doc:detail", args=(doc_id,)))

    try:
        doc = Doc.objects.get(pk=doc_id)
    except Doc.DoesNotExist as e:
        return TemplateResponse(request, "doc.html", {
            'page_title': f"Documentation - ID {doc_id}",
            'doc_id': doc_id,
            'error': 404,
        })

    trForm = TimeRecordForm(request.POST, user=request.user)
    if trForm.is_valid():
        time_record = trForm.save(commit=False)
        time_record.doc = doc
        time_record.save()
        doc.save()

    return HttpResponseRedirect(reverse("doc:detail", args=(doc_id,)))


@login_required
@permission_required("doc.time_records_set_settled")
def time_records_set_settled(request, **kwargs):
    """
    Sets the selected time records to settled.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("doc:home"))

    success = False
    error = False

    try:
        trs = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user)
        for tr in trs:
            tr.settled_at = timezone.now()
            tr.save()
            tr.doc.save()
            success = [
                'The selected entries were set to settled.',
                'Please use the back button in the toolbar to return to the start page.',
            ]
    except Exception as e:
        error = [e]

    return TemplateResponse(request, "message.html", {
        'page_title': f"Set Time Records To Settled",
        'success': success,
        'error': error,
    })


@login_required
@permission_required("doc.time_records_set_deleted")
def time_records_set_deleted(request, **kwargs):
    """
    Sets the selected time records to deleted.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("doc:home"))

    success = False
    error = False

    try:
        trs = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user)
        for tr in trs:
            tr.deleted_at = timezone.now()
            tr.save()
            tr.doc.save()
            success = [
                'The selected entries have been deleted.',
                'Please use the back button in the toolbar to return to the start page.',
            ]
    except Exception as e:
        error = [e]

    return TemplateResponse(request, "message.html", {
        'page_title': f"Delete Time Records",
        'success': success,
        'error': error,
    })


@login_required
@permission_required("doc.time_analyse")
def time_analyse_download(request, **kwargs):
    """
    Returns a download in the specified format.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("doc:home"))

    rows = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user, ) \
        .order_by('date', 'doc__title', 'description') \
        .values_list('date', 'doc__title', 'description', 'time', named=True)

    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"timereport_{timestamp}"

    if request.POST.get('dcsv', default=False):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Item", "Task", "Time [h]"])
        for row in rows:
            writer.writerow([row.date, row.doc__title, row.description, row.time])
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response
    elif request.POST.get('dxlsx', default=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Time Report"
        headers = ["Date", "Item", "Task", "Time [h]"]
        ws.append(headers)

        # Freeze the first row and set it as title row
        ws.freeze_panes = "A2"
        ws.print_title_rows = '1:1'

        # Header in bold letters
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Add rows
        for row in rows:
            ws.append([row.date, row.doc__title, row.description, row.time])

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 11

        # Row height
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 31

        # Alignment of the columns A, B and C
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Allow text wrap in the columns B and C
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        # Alignment of the column D
        ws["D1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Number format and alignment of the column D
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        thin_border = Border(bottom=Side(style="thin", color="000000"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

        # Page properties
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.59
        ws.page_margins.right = 0.59
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response
    elif request.POST.get('dpdf', default=False):
        """The PDF format will not be implemented in the near future."""
        total = sum((Decimal(str(row.time)) for row in rows), Decimal('0.0'))

    response = HttpResponse("No format specified", content_type="text/plain")
    response["Content-Disposition"] = f'attachment; filename="{filename}.txt"'
    return response


@login_required
@permission_required("doc.time_analyse")
def time_analyse_preview(request, **kwargs):
    """
    Lists the selected time records.
    """

    if not request.POST:
        return HttpResponseRedirect(reverse("doc:home"))

    rows = TimeRecord.objects.filter(id__in=request.POST.getlist('tr') or [], user=request.user, ) \
        .order_by('date', 'doc__title', 'description') \
        .values_list('date', 'doc__title', 'description', 'time', named=True)

    total = sum((Decimal(str(row.time)) for row in rows), Decimal('0.0'))

    return TemplateResponse(request, "tr_analyse_preview.html", {
        'page_title': f"Time Records Selection",
        'rows': rows,
        'total': total,
    })


@login_required
@permission_required("doc.time_analyse")
def time_analyse_selection(request, **kwargs):
    """
    Lists all unsettled time record entries of the given docs and allows to discriminate
    to a certain period of time and to select a set of entries to be previewed, downloaded
    and set to billed.
    """

    entities = TimeRecord.objects.filter(
        deleted_at__isnull=True,
        settled_at__isnull=True,
        doc_id__in=request.GET.getlist('doc') or [],
        user=request.user,
    )

    date_range = entities.aggregate(
        min_date=Min('date'),
        max_date=Max('date')
    )

    dgte = parse_valid_date(request.GET.get('dgte'), date_range['min_date'])
    dlte = parse_valid_date(request.GET.get('dlte'), date_range['max_date'])
    if dgte and dlte and dgte > dlte:
        dgte, dlte = dlte, dgte

    entities = entities.filter(date__gte=dgte, date__lte=dlte).order_by('date', 'doc__title', 'description')

    return TemplateResponse(request, "tr_analyse_selection.html", {
        'page_title': f"Time Records Selection",
        'min_date': date_range['min_date'],
        'max_date': date_range['max_date'],
        'dgte': dgte,
        'dlte': dlte,
        'docs': request.GET.getlist('doc') or [],
        'entities': entities,
    })
